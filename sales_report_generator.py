# Sales report generator
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from pathlib import Path
import openpyxl, sys
from openpyxl.utils import get_column_letter
print('=' * 35)
print('Sales Report Generator')
print('=' * 35)
while True:
    user_file_path = Path(input('Enter excel file path: ').strip())
    if user_file_path.is_file():
        break
    print('Invalid file path. please try again')
# make reports folder
sales_report_folder = user_file_path.parent / 'Reports'
sales_report_path = sales_report_folder / f'{user_file_path.stem}_report.xlsx'
sales_report_folder.mkdir(parents = True, exist_ok = True)
try:
    wb = openpyxl.load_workbook(user_file_path)
except Exception:
    sys.exit('Cannot open workbook!')
while True:
    print(f'Available sheets: {wb.sheetnames}')
    print('choose sheet:(Case sensitive) ')
    user_sheet = input('>').strip()
    if user_sheet not in wb.sheetnames:
        print('Selected worksheet not found!')
        continue
    sheet = wb[user_sheet]
    if sheet.max_row == 1 and sheet.max_column == 1 and sheet['A1'].value is None:
        print('Selected sheet is empty. Exiting...')
        sys.exit()
    break
price_col = None
units_col = None
product_col = None
category_col = None
# find the columns
for col in range(1, sheet.max_column +1):
    header = str(sheet.cell(row = 1 , column = col).value).strip().lower()
    if header == 'units':
        units_col = col
    elif header == 'price':
        price_col = col
    elif header == 'product':
        product_col = col
    elif header == 'category':
        category_col = col
# Validate after searching all headers
if price_col is None or units_col is None:
    sys.exit('Price or Units column not found')
if product_col is None:
    sys.exit('Product column not found!')
if category_col is None:
    sys.exit('Category column not found.')
# calculate revenue and find best-selling product
product_units = {}
total_revenue = 0
product_revenue = {}
total_price = 0
product_count = 0
category_revenue ={}
for row in range(2, sheet.max_row +1):
    units = sheet.cell(row = row, column = units_col).value
    price = sheet.cell(row = row, column = price_col).value 
    product = sheet.cell(row = row, column = product_col).value
    category = sheet.cell(row = row, column = category_col).value
    if (price is None 
        or units is None 
        or not isinstance(price, (int, float)) 
        or not isinstance(units, (int, float))):
        continue
    if product is None:
        continue
    revenue = units * price 
    if product in product_revenue:
        product_revenue[product] += revenue
    else:
        product_revenue[product] = revenue
    total_revenue += revenue
    if product in product_units:
        product_units[product] += units
    else:
        product_units[product] = units
    total_price += price
    product_count += 1
    if category is None :
        continue
    if category in category_revenue:
        category_revenue[category] += revenue
    else:
        category_revenue[category] = revenue
total_units = sum(product_units.values())
#average product price
average_product_price = 0
if product_count > 0:
    average_product_price = total_price / product_count
else:
    average_product_price = 0
# find the best seller
highest_unit = 0
best_selling = None
for product, unit in product_units.items():
    if unit > highest_unit:
        highest_unit = unit
        best_selling = product
# find the highest revenue
highest_revenue = 0
highest_revenue_product = None
for product, revenue in product_revenue.items():
    if revenue > highest_revenue:
        highest_revenue = revenue
        highest_revenue_product = product
# calculate highest revenue category
best_category = None
best_category_revenue = 0
for category, revenue in category_revenue.items():
    if revenue > best_category_revenue:
        best_category_revenue = revenue
        best_category = category
print('\nSummary')
print(f'\ntotal revenue: {total_revenue}')
print(f'\nTotal units sold: {total_units}')
print(f'\nBest-selling product: {best_selling}')
print(f'\nHighest revenue product: {highest_revenue_product}')
print(f'\nAverage product price: {average_product_price}')
print(f'\nHighest revenue category: \n{best_category} : {best_category_revenue}')
# creat report sheet
report_sheet = wb.create_sheet('Sales Report')
report_sheet['A1'] = 'Sales Summary'

report_sheet['A3'] = 'Metric'
report_sheet['B3'] = 'Value'

report_sheet['A4'] = 'Total Revenue'
report_sheet['B4'] = total_revenue

report_sheet['A5'] = 'Total Units Sold'
report_sheet['B5'] = total_units

report_sheet['A6'] = 'Best Seller'
report_sheet['B6'] = best_selling

report_sheet['A7'] = 'Highest Revenue Product'
report_sheet['B7'] = highest_revenue_product

report_sheet['A8'] = 'Highest Revenue Category'
report_sheet['B8'] = best_category

report_sheet['A9'] = 'Average Product Price'
report_sheet['B9'] = average_product_price

report_sheet['A12'] = 'Revenue by Category'
report_sheet['A14'] = 'Category'
report_sheet['B14'] = 'Revenue'
row = 15
for category, revenue in category_revenue.items():
    report_sheet.cell(row = row, column = 1).value = category
    report_sheet.cell(row = row, column = 2).value = revenue
    row += 1
# format the report sheet
header_font = Font(bold = True, size = 14, color = '00FFFFFF')
header_fill = PatternFill(fill_type = 'solid', end_color = '1F4E79')
text_alignment = Alignment(horizontal = 'center', vertical = 'center')
right_border = Side(border_style = 'thick', color = '7F7F7F')
col_right_border = Border(right = right_border)
for col in range(1, report_sheet.max_column + 1):
    max_length = 0
    for row in range(1, report_sheet.max_row + 1):
        cell = report_sheet.cell(row = row, column = col)
        cell.alignment = text_alignment
        cell.border = col_right_border
        value = cell.value
        if value is not None:
            max_length = max(max_length, len(str(value)))
            letter = get_column_letter(col)
            report_sheet.column_dimensions[letter].width = max_length + 5
for col in range(1, report_sheet.max_column + 1):
    report_sheet.cell(row =1, column =col).font = header_font
    report_sheet.cell(row =1, column =col).fill = header_fill
for cell in ('A3','B3','A12', 'A14', 'B14' ):
    report_sheet[cell].font = header_font
    report_sheet[cell].fill = header_fill
report_sheet['B4'].number_format = '$#,##0.00'
report_sheet['B9'].number_format = '$#,##0.00'
for row in range(15, report_sheet.max_row + 1):
    report_sheet[f'B{row}'].number_format = '$#,##0.00'
report_sheet.freeze_panes = 'A4'
wb.save(sales_report_path)
wb.close()
print('Report sheet successfully created.')
print(f'\nSaved at: {sales_report_path}')