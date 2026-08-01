# Sales report generator
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from pathlib import Path
import openpyxl, sys, argparse, logging
from openpyxl.utils import get_column_letter
print('=' * 35)
print('Sales Report Generator')
print('=' * 35)
logging.basicConfig(
    filename = 'sales_report.log', 
    level= logging.INFO, 
    format = '%(asctime)s - %(levelname)s - %(message)s'
    )
def get_arguments():
    parser = argparse.ArgumentParser(
        description = 'Generate sales report from Excel files'
        )
    parser.add_argument(
        'file',
        help = 'Excel file to process.',
        type = Path
        )
    parser.add_argument(
        '--sheet', 
        help = 'The exact worksheet name to process.(optional)'
        )
    parser.add_argument(
        '--output', 
        type = Path, 
        help = 'Output report file path.'
        )
    args = parser.parse_args() 
    return args
args = get_arguments()
if not args.file.is_file():
    sys.exit(' File not found')
user_file_path = args.file
user_sheet = args.sheet.strip() if args.sheet else None
logging.info(f'Processing file: {user_file_path}')
def open_wb(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        return workbook
    except PermissionError:
        logging.error('Workbook is open in another program.')
        sys.exit('File is open in Excel. Close it and try again.')
    except Exception as e:
        logging.error(f'Cannot open file: {e}')
        sys.exit(f'Cannot open file: {e}')
wb = open_wb(user_file_path)
logging.info('Workbook open successfully.')
def get_sheet(workbook, sheet_name):
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            logging.error('Sheet name does not exist.')
            sys.exit('Worksheet not found') 
        return workbook[sheet_name]
    else:
        return workbook.active
sheet = get_sheet(wb, user_sheet)
logging.info(f'Processing worksheet: {sheet.title}')
if sheet.max_row == 1 and sheet.max_column == 1 and sheet['A1'].value is None:
    logging.error('Worksheet is empty!')
    sys.exit('Selected sheet is empty. Exiting...')
# find the columns
logging.info('Processing columns...')
def find_columns(sheet):
    price_col = None
    units_col = None
    product_col = None
    category_col = None
    for col in range(1, sheet.max_column +1):
        header = str(
            sheet.cell(row = 1 , column = col).value
            ).strip().lower()
        if header == 'units':
            units_col = col
        elif header == 'price':
            price_col = col
        elif header == 'product':
            product_col = col
        elif header == 'category':
            category_col = col
    # Validate after searching all headers
    if price_col is None or units_col is None:
        logging.error('Price or Units column not found')
        sys.exit('Price or Units column not found')
    if product_col is None:
        logging.error('Product column not found!')
        sys.exit('Product column not found!')
    if category_col is None:
        logging.error('Category column not found.')
        sys.exit('Category column not found.')
    return {
        'units': units_col, 
        'price' : price_col, 
        'product' : product_col, 
        'category' :category_col
        }
columns = find_columns(sheet)
logging.info(f'Columns detected: {columns}')
# calculate revenue and find best-selling product
logging.info('Calculating sales...')
def calculate_sales(sheet, columns):
    product_units = {}
    total_revenue = 0
    product_revenue = {}
    total_price = 0
    product_count = 0
    category_revenue ={}
    for row in range(2, sheet.max_row +1):
        units = sheet.cell(row = row, column = columns['units']).value
        price = sheet.cell(row = row, column = columns['price']).value 
        product = sheet.cell(row = row, column = columns['product']).value
        category = sheet.cell(row = row, column = columns['category']).value
        if (price is None 
            or units is None 
            or not isinstance(price, (int, float)) 
            or not isinstance(units, (int, float))):
                logging.warning(f'Skipped invalid data at row: {row}')
                continue
        if product is None:
            logging.warning(f"Skipped row {row}: missing product name.")
            continue
        revenue = units * price 
        if product in product_revenue:
            product_revenue[product] += revenue
        else:
            product_revenue[product] = revenue
        total_revenue += revenue
        if product in product_units:
            product_units[product] += units
        else:
            product_units[product] = units
        total_price += price
        product_count += 1
        if category is not None :
            if category in category_revenue:
                category_revenue[category] += revenue
            else:
                category_revenue[category] = revenue
    return {
        'product_units': product_units, 
        'product_revenue': product_revenue, 
        'category_revenue': category_revenue, 
        'total_revenue': total_revenue, 
        'total_price': total_price, 
        'product_count': product_count
        }
sales_data = calculate_sales(sheet, columns)
total_units = sum(sales_data['product_units'].values())
logging.info('Sales calculations completed.')
logging.info(f"Total revenue: {sales_data['total_revenue']}")
logging.info(f"Total units sold: {total_units}")
#average product price
average_product_price = 0
if sales_data['product_count'] > 0:
    average_product_price = sales_data['total_price'] / sales_data['product_count']
else:
    average_product_price = 0
# find the best seller
highest_unit = 0
best_selling = None
for product, unit in sales_data['product_units'].items():
    if unit > highest_unit:
        highest_unit = unit
        best_selling = product
logging.info(f'Best seller: {best_selling}')
# find the highest revenue
highest_revenue = 0
highest_revenue_product = None
for product, revenue in sales_data['product_revenue'].items():
    if revenue > highest_revenue:
        highest_revenue = revenue
        highest_revenue_product = product
# calculate highest revenue category
best_category = None
best_category_revenue = 0
for category, revenue in sales_data['category_revenue'].items():
    if revenue > best_category_revenue:
        best_category_revenue = revenue
        best_category = category
# make reports folder
sales_report_folder = user_file_path.parent / 'Reports'
sales_report_folder.mkdir(parents = True, exist_ok = True)
print('\nSummary')
print(f"\ntotal revenue: {sales_data['total_revenue']}")
print(f'\nTotal units sold: {total_units}')
print(f'\nBest-selling product: {best_selling}')
print(f'\nHighest revenue product: {highest_revenue_product}')
print(f'\nAverage product price: {average_product_price}')
print(f'\nHighest revenue category: \n{best_category} : {best_category_revenue}')
# create report sheet
logging.info('Creating report sheet...')
analysis = {
    'sales_data': sales_data,
    'total_units': total_units,
    'best_selling': best_selling,
    'highest_revenue_product': highest_revenue_product,
    'best_category': best_category,
    'average_product_price': average_product_price
    }
def create_report(workbook, analysis):
    if 'Sales Report' in workbook.sheetnames:
        del workbook['Sales Report']
    report_sheet = workbook.create_sheet('Sales Report')
    report_sheet['A1'] = 'Sales Summary'

    report_sheet['A3'] = 'Metric'
    report_sheet['B3'] = 'Value'

    report_sheet['A4'] = 'Total Revenue'
    report_sheet['B4'] = analysis['sales_data']['total_revenue']

    report_sheet['A5'] = 'Total Units Sold'
    report_sheet['B5'] = analysis['total_units']

    report_sheet['A6'] = 'Best Seller'
    report_sheet['B6'] = analysis['best_selling']

    report_sheet['A7'] = 'Highest Revenue Product'
    report_sheet['B7'] = analysis['highest_revenue_product']

    report_sheet['A8'] = 'Highest Revenue Category'
    report_sheet['B8'] = analysis['best_category']

    report_sheet['A9'] = 'Average Product Price'
    report_sheet['B9'] = analysis['average_product_price']

    report_sheet['A12'] = 'Revenue by Category'
    report_sheet['A14'] = 'Category'
    report_sheet['B14'] = 'Revenue'
    row = 15
    for category, revenue in analysis['sales_data']['category_revenue'].items():
        report_sheet.cell(row = row, column = 1).value = category
        report_sheet.cell(row = row, column = 2).value = revenue
        row += 1
    return report_sheet
report_sheet = create_report(wb,analysis)
logging.info('Report sheet successfully created.')
# format the report sheet
logging.info('Formatting the report sheet.')
header_font = Font(bold = True, size = 14, color = '00FFFFFF')
header_fill = PatternFill(fill_type = 'solid', fgColor = '1F4E79')
text_alignment = Alignment(horizontal = 'center', vertical = 'center')
right_border = Side(border_style = 'thick', color = '7F7F7F')
col_right_border = Border(right = right_border)
formats = {
    'header_font': header_font,
    'header_fill': header_fill,
    'text_alignment': text_alignment,
    'col_right_border': col_right_border
    }
def format_sheet(report_sheet, formats):
    for col in range(1, report_sheet.max_column + 1):
        max_length = 0
        for row in range(1, report_sheet.max_row + 1):
            cell = report_sheet.cell(row = row, column = col)
            cell.alignment = formats['text_alignment']
            cell.border = formats['col_right_border']
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))
            letter = get_column_letter(col)
            report_sheet.column_dimensions[letter].width = max_length + 5
    for col in range(1, report_sheet.max_column + 1):
        report_sheet.cell(row =1, column =col).font = formats['header_font']
        report_sheet.cell(row =1, column =col).fill = formats['header_fill']
    for cell in ('A3','B3','A12', 'A14', 'B14' ):
        report_sheet[cell].font = formats['header_font']
        report_sheet[cell].fill = formats['header_fill']
    report_sheet['B4'].number_format = '$#,##0.00'
    report_sheet['B9'].number_format = '$#,##0.00'
    for row in range(15, report_sheet.max_row + 1):
        report_sheet[f'B{row}'].number_format = '$#,##0.00'
    report_sheet.freeze_panes = 'A4'
format_sheet(report_sheet, formats)
if args.output is not None:
    args.output.parent.mkdir(parents = True, exist_ok = True)
    sales_report_path = args.output
else:
    sales_report_path = sales_report_folder / f'{user_file_path.stem}_report.xlsx'
wb.save(sales_report_path)
wb.close()
logging.info(f'Report saved successfully: {sales_report_path}')
print(f'\nSaved at: {sales_report_path}')